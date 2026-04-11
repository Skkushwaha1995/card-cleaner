import streamlit as st
import pandas as pd
import re
from io import BytesIO

st.set_page_config(page_title="CarDekho Data Cleaner", page_icon="🚗", layout="wide")

st.title("🚗 CarDekho Data Cleaner")
st.caption("Console se copy kiya data paste karo — auto clean + merge + Excel download")

# ─── Helper Functions ───────────────────────────────────────────────────────

SKIP_PATTERNS = [
    r"^view\s*(more|all|less)$",
    r"^show\s*(more|all|less)$",
    r"^hide\s*(common|features)$",
    r"^\s*$",
    r"^space image$",
    r"^ad$",
    r"^get emi offers$",
    r"^based on\s*\d+\s*reviews?$",
    r"^\d+\s*reviews?$",
    r"^all\s+.*cars?$",
    r"^view\s+.*colours?$",
    r"^view\s+.*offers?$",
]

def should_skip(val: str) -> bool:
    v = val.strip().lower()
    # Skip if matches any pattern
    if any(re.match(p, v) for p in SKIP_PATTERNS):
        return True
    # Skip cells that contain "based on" + number (review counts embedded in value)
    if re.search(r"based on\s*\d+\s*reviews?", v):
        return True
    # Skip standalone rating numbers like "4.7"
    if re.match(r"^\d\.\d$", v):
        return True
    return False

def clean_val(v: str) -> str:
    v = v.strip()
    if re.match(r"^(yes|✓|✔|☑|✅|\u2713|\u2714|\u2705)$", v, re.IGNORECASE):
        return "Yes"
    if re.match(r"^(no|✗|✘|☒|❌|\u2717|\u2718|\u274c)$", v, re.IGNORECASE):
        return "No"
    if re.match(r"^(-+|–+|—+|na|n/a|not available|not applicable)$", v, re.IGNORECASE):
        return "N/A"
    v = re.sub(r"\s*space image\s*", "", v, flags=re.IGNORECASE)
    # Clean price — keep only first line (remove EMI/extra text)
    if re.match(r"^rs", v, re.IGNORECASE):
        v = v.split("\n")[0].strip()
    v = v.rstrip("*").strip()
    if not v:
        return "N/A"
    return v

def parse_clipboard(raw: str):
    """Parse pipe-separated clipboard text into headers + rows with sections."""
    lines = [l.strip() for l in raw.split("\n") if l.strip()]
    headers = []
    rows = []
    current_section = ""

    for line in lines:
        if should_skip(line):
            continue
        cols = [c.strip() for c in line.split("|")]

        # Skip if first cell is a skip pattern
        if should_skip(cols[0]):
            continue

        rest = [clean_val(c) for c in cols[1:]]

        # Detect section header: all other cols empty
        if len(cols) >= 2 and all(c.strip() == "" for c in cols[1:]):
            current_section = clean_val(cols[0])
            continue

        # Single column line → section header
        if len(cols) == 1:
            current_section = clean_val(cols[0])
            continue

        # First data row → treat as header if headers not set
        if not headers:
            headers = [clean_val(c) for c in cols]
            if headers[0] == "":
                headers[0] = "Field"
            continue

        rows.append({
            "Section": current_section,
            "values": [clean_val(cols[0])] + rest
        })

    return headers, rows

def build_dataframe(headers, rows):
    """Convert parsed data into a DataFrame."""
    records = []
    for row in rows:
        record = {"Section": row["Section"]}
        for i, h in enumerate(headers):
            record[h] = row["values"][i] if i < len(row["values"]) else ""
        records.append(record)
    df = pd.DataFrame(records)
    return df

def merge_dataframes(dfs: list) -> pd.DataFrame:
    """Merge multiple DataFrames column-wise on 'Field' + 'Section'."""
    if not dfs:
        return pd.DataFrame()
    if len(dfs) == 1:
        return dfs[0]

    # Use first df's Field column as base
    field_col = dfs[0].columns[1]  # First col after Section
    base = dfs[0].copy()

    for i, df in enumerate(dfs[1:], start=2):
        # Get car columns (everything except Section and Field)
        car_cols = [c for c in df.columns if c not in ["Section", field_col]]
        # Rename to avoid collision
        rename_map = {c: f"{c} ({i})" for c in car_cols if c in base.columns}
        df = df.rename(columns=rename_map)
        car_cols_renamed = [rename_map.get(c, c) for c in car_cols]
        merge_df = df[["Section", field_col] + car_cols_renamed]
        base = pd.merge(base, merge_df, on=["Section", field_col], how="outer")

    return base

def to_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CarDekho")
        ws = writer.sheets["CarDekho"]

        # Style header row
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="1F4E79")
        section_fill = PatternFill("solid", fgColor="D6E4F0")
        thin = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin

        # Style data rows
        section_col_idx = None
        for idx, col in enumerate(df.columns, 1):
            if col == "Section":
                section_col_idx = idx

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            section_val = ""
            if section_col_idx:
                section_val = ws.cell(row=row_idx, column=section_col_idx).value or ""

            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.column == section_col_idx and section_val:
                    cell.fill = section_fill
                    cell.font = Font(bold=True, color="1F4E79", size=10)
                elif str(cell.value) == "Yes":
                    cell.font = Font(color="276221", bold=True)
                elif str(cell.value) == "No":
                    cell.font = Font(color="9C0006", bold=True)
                elif str(cell.value) == "N/A":
                    cell.font = Font(color="888888", italic=True)
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "B2"

    return output.getvalue()

# ─── Session State ────────────────────────────────────────────────────────────
if "clipboards" not in st.session_state:
    st.session_state.clipboards = [""]
if "merged_df" not in st.session_state:
    st.session_state.merged_df = None

# ─── UI ───────────────────────────────────────────────────────────────────────

col1, col2 = st.columns([1, 1])
with col1:
    if st.button("➕ Add Another Clipboard", help="Ek aur data block paste karne ke liye"):
        st.session_state.clipboards.append("")

with col2:
    if len(st.session_state.clipboards) > 1:
        if st.button("➖ Remove Last Clipboard"):
            st.session_state.clipboards.pop()
            st.rerun()

st.divider()

# ─── Clipboard Text Areas ─────────────────────────────────────────────────────
for i in range(len(st.session_state.clipboards)):
    label = f"📋 Clipboard {i+1}" if len(st.session_state.clipboards) > 1 else "📋 Data Paste Karo"
    hint = "Pehle car ka data yahan paste karo (console se copy kiya)" if i == 0 else f"Doosre/teesre car ka data yahan paste karo — columns merge ho jayenge"
    st.session_state.clipboards[i] = st.text_area(
        label,
        value=st.session_state.clipboards[i],
        height=180,
        placeholder=hint,
        key=f"cb_{i}"
    )

st.divider()

# ─── Process Button ───────────────────────────────────────────────────────────
if st.button("🧹 Clean & Merge", type="primary", use_container_width=True):
    all_dfs = []
    errors = []

    for i, raw in enumerate(st.session_state.clipboards):
        if not raw.strip():
            continue
        try:
            headers, rows = parse_clipboard(raw)
            if not rows:
                errors.append(f"Clipboard {i+1}: Koi data parse nahi hua.")
                continue
            df = build_dataframe(headers, rows)
            all_dfs.append(df)
        except Exception as e:
            errors.append(f"Clipboard {i+1} error: {e}")

    if errors:
        for e in errors:
            st.warning(e)

    if all_dfs:
        if len(all_dfs) > 1:
            merged = merge_dataframes(all_dfs)
            st.success(f"✅ {len(all_dfs)} clipboards merge ho gaye — {len(merged)} rows, {len(merged.columns)} columns!")
        else:
            merged = all_dfs[0]
            st.success(f"✅ {len(merged)} rows clean ho gaye!")
        st.session_state.merged_df = merged
    else:
        st.error("Koi valid data nahi mila. Kuch paste karo pehle.")

# ─── Preview & Download ───────────────────────────────────────────────────────
if st.session_state.merged_df is not None:
    df = st.session_state.merged_df

    # Stats
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Fields", len(df))
    m2.metric("Columns", len(df.columns))
    sections = df["Section"].nunique() if "Section" in df.columns else 0
    m3.metric("Sections", sections)
    cars = len([c for c in df.columns if c not in ["Section"]]) - 1
    m4.metric("Cars", max(cars, 0))

    st.subheader("📊 Preview")

    # Style the dataframe
    def highlight(row):
        styles = []
        for col in row.index:
            val = str(row[col])
            if col == "Section":
                styles.append("background-color: #D6E4F0; font-weight: bold; color: #1F4E79")
            elif val == "Yes":
                styles.append("color: #276221; font-weight: bold")
            elif val == "No":
                styles.append("color: #9C0006; font-weight: bold")
            elif val == "N/A":
                styles.append("color: #888888; font-style: italic")
            else:
                styles.append("")
        return styles

    st.dataframe(
        df.style.apply(highlight, axis=1),
        use_container_width=True,
        height=400
    )

    # Download buttons
    d1, d2 = st.columns(2)
    with d1:
        excel_data = to_excel(df)
        st.download_button(
            "⬇️ Download Excel (.xlsx)",
            data=excel_data,
            file_name="cardekho_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with d2:
        csv_data = df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            "⬇️ Download CSV",
            data=csv_data,
            file_name="cardekho_data.csv",
            mime="text/csv",
            use_container_width=True
        )

# ─── Instructions ─────────────────────────────────────────────────────────────
with st.expander("📖 Kaise Use Karein?"):
    st.markdown("""
**Step 1 — CarDekho page kholo**
- Browser mein `F12` dabao → Console tab

**Step 2 — Pasting enable karo**
- Console mein type karo: `allow pasting` → Enter

**Step 3 — Ye code paste karo aur Enter dabao:**
```javascript
let rows = document.querySelectorAll("table tr");
let result = "";
rows.forEach(row => {
    let cells = row.querySelectorAll("td, th");
    let line = Array.from(cells).map(c => c.innerText.trim()).join(" | ");
    if(line.trim()) result += line + "\\n";
});
copy(result);
alert("Copy ho gaya!");
```

**Step 4 — Yahan paste karo**
- Copied data ऊपर textarea mein paste karo

**Multiple Cars ke liye:**
- `➕ Add Another Clipboard` dabao
- Doosre car ka data doosre textarea mein paste karo
- `Clean & Merge` dabao — sab columns merge ho jayenge!

**Auto-cleaned cheezein:**
- ✓ / ✗ → Yes / No
- `View More`, `View All` rows remove
- `*` aur extra spaces clean
- Section headers alag dikhenge
    """)
